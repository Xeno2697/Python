from board import board
from Qplayer import Qplayer
from easyAI import eAI
from tqdm import tqdm
import openpyxl 

class main:
    def __init__(self,X,Y,winnum) -> None:
        self.X = X
        self.Y = Y
        self.board = board(X,Y,winnum)
        self.Qplayer = Qplayer(self.board,0.35,0.95,0.0)
        self.easyai = eAI()
        return

    def start(self,time):
        qwin = 0
        ewin = 0

        for i in tqdm(range(time)):
            #EvsQ
            self.board.reset()
            for j in range((int)((self.X*self.Y)/2)):
                c = self.easyai.choice(self.board,False)
                if c>=10000:
                    self.Qplayer.learn(self.board,c%10000,False,1)
                    ewin += 1
                    break
                else:
                    self.Qplayer.learn(self.board,c%10000,False,0)
                    self.board.push(c,1)
                c = self.Qplayer.choice(self.board,True)
                if c>=10000:
                    self.Qplayer.learn(self.board,c%10000,True,1)
                    qwin += 1
                    break
                else:
                    self.Qplayer.learn(self.board,c%10000,True,0)
                    self.board.push(c,-1)

            #QvsE
            self.board.reset()
            for j in range((int)((self.X*self.Y)/2)):
                c = self.Qplayer.choice(self.board,False)
                if c>=10000:
                    self.Qplayer.learn(self.board,c%10000,False,1)
                    qwin += 1
                    break
                else:
                    self.Qplayer.learn(self.board,c%10000,False,0)
                    self.board.push(c,1)
                c = self.easyai.choice(self.board,True)
                if c>=10000:
                    self.Qplayer.learn(self.board,c%10000,True,1)
                    ewin += 1
                    break
                else:
                    self.Qplayer.learn(self.board,c%10000,True,0)
                    self.board.push(c,-1)
            
            #QvsQ
            self.board.reset()
            for j in range(int((self.X*self.Y)/2)):
                c = self.Qplayer.choice(self.board,False)
                if c>=10000:
                    self.Qplayer.learn(self.board,c%10000,False,1)
                    break
                else:
                    self.Qplayer.learn(self.board,c,False,0)
                    self.board.push(c,1)
                c = self.Qplayer.choice(self.board,True)
                if c>=10000:
                    self.Qplayer.learn(self.board,c%10000,True,1)
                    break
                else:
                    self.Qplayer.learn(self.board,c,True,0)
                    self.board.push(c,-1)
            #self.board.print()
        
        print("EasyAI win ",end="")
        print(ewin)
        print("Qlearning win ",end="")
        print(qwin)

        wb = openpyxl.Workbook()
        s1 = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for i,key in enumerate(self.Qplayer.Q):
            s1.cell(row=i+1,column=1,value="Qarray")
            for j in range(self.X):
                s1.cell(row=i+1,column=j+2,value=self.Qplayer.Q[key][j])
        wb.save('test.xlsx')
        return